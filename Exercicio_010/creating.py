# pip install selenium
# pip openpyxl selenium

from pathlib import Path
from time import sleep

from selenium import webdriver
from selenium.webdriver.chrome.service import Service

from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent

CHROME_DRIVER_PATH = ROOT_FOLDER / 'drivers' / 'chromedriver.exe'

WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

# picking up a table with dollar values

chorme_service = Service(executable_path = str(CHROME_DRIVER_PATH),)

chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument('--headless')

browser = webdriver.Chrome(
    service=chorme_service,
    options= chrome_options
)

browser.get('https://www.x-rates.com/table/?from=USD&amount=1')
tables = browser.find_elements(By.TAG_NAME, 'table')

table_alphabetical_order = (tables[1].text).split('\n')

# organizing the table

countries_currencies = []

for single_country in table_alphabetical_order:

    country_data = []

    value = single_country[::-1]

    first_division = value.index(' ')
    second_division = value.index(' ', int(first_division)+1)

    usd_inv = value[0:first_division]
    usd_inv = usd_inv[::-1]

    usd_value = value[first_division+1:second_division]
    usd_value = usd_value[::-1]

    country_currency = value[second_division+1:]
    country_currency = country_currency[::-1]

    country_data.append(country_currency)
    country_data.append(usd_value)

    countries_currencies.append(country_data)

# Delete header

del countries_currencies[0]

#Creating a SpreadSheet

workbook = Workbook()

worksheet:Worksheet = workbook.active   #type: ignore

worksheet.cell(1, 1, 'Currency')
worksheet.cell(1, 2, 'Value')

for moeda in countries_currencies:
    worksheet.append(moeda)

workbook.save(WORKBOOK_PATH)
