from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

import os



ROOT_FOLDER = Path(__file__).parent

WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'
WORKBOOK_CREATING_PATH = ROOT_FOLDER / 'creating.py'

def open_country():
    try:
        workbook: Workbook = load_workbook(WORKBOOK_PATH)

    except:
        os.system(f'python {WORKBOOK_CREATING_PATH}')
        workbook: Workbook = load_workbook(WORKBOOK_PATH)

    finally:
        worksheet:Worksheet = workbook.active   #type: ignore

    country_currency = []

    for row in worksheet.iter_rows(min_row=2):
        informations = []
        for cell in row:
            informations.append(cell.value)
        country_currency.append(informations)
    return country_currency